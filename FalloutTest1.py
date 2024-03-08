

"""This is an earlier proof of concept for the main.py program
   It is not well written, but allowed me to gain some experience 
   with Pandas"""



import time
from pynput import keyboard
import pandas as pd
import numpy as np
import openpyxl as xl
from datetime import datetime
import os

StartCount = 0
CountCounter = 0
green = 0
blue = 0
yellow = 0
forrest = 0
red = 0
total = 0
data = np.array([0,0,0,0,0,0,0,0,0])
now = datetime.now()
brown = 0
#current_time = now.strftime("%Y-%m-%d %H:%M:%S")

def on_press(key):
    global green, blue, yellow, forrest, red, brown, total
    if key == keyboard.Key.esc:
        return False #stop-listener
    try:
        k = key.char
    except:
        k = key.name
    if k in ['s']:
        print ('You have pressed "s" to start your run')
        print('Good Luck!')
        update_counter_start()
        if get_start_count() == 0:
            print ('Start time is:', get_local_time())
            start_update()
            print ('YOU ARE STARTING')
    if k in ['g']:
        green +=1
    if k in ['z']:
        brown +=1
    if k in ['b']:
        blue +=1
    if k in ['y']:
        yellow +=1
    if k in ['f']:
        forrest +=1
    if k in ['t']:
        total +=1
    if k in ['r']:
        red +=1
    if k in ['c']:
        if CountCounter > 0:
            print ('You have logged an attempt')
            update_array(brown, green, blue, yellow, forrest, red, total) 
            green = 0
            brown = 0 
            blue = 0
            yellow = 0
            forrest = 0
            red = 0
            total = 0 
        else:
            print ("\nPlease press 's' to begin.")
    if k in ('2'):
        end_update_counter ()
    '''else:
        if CountCounter >0:
            print ("\nPlease select one of the available options")
        else:
            print ("\nPlease press 's' to begin.")'''


def get_start_count():
    return StartCount

def start_update():
    global StartCount
    StartCount = StartCount + 1

def get_seconds():
    seconds = time.time()
    return seconds

def get_local_time():
    local_time = time.ctime(get_seconds())
    return local_time

def update_counter_start():
    global CountCounter
    global data
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d %H:%M:%S")
    if CountCounter == 0:
        newrow = [0,current_time,0,0,0,0,0,0,0]
        data = np.vstack([newrow])
        CountCounter +=1

def update_array (Brown, Green, Blue, Yellow, Forrest, Red, Total):
    global CountCounter
    global data
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d %H:%M:%S")
    if CountCounter > 0:
        newrow = [CountCounter,current_time,Brown,Green,Blue,Yellow,Forrest,Red,Total]
        data = np.vstack([data, newrow])
        CountCounter +=1

    print (data)

def end_update_counter ():
    global CountCounter
    global data
    count2 = 0
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d %H:%M:%S")
    newrow = [0,current_time,0,0,0,0,0,0,0]
    data = np.vstack([data, newrow])
    newrow = ['Total Attempts','Total Time','Total Brown','Total Green','Total Blue','Total Yellow','Total Forrest','Total Red','Total Spawns']
    data = np.vstack([data, newrow])
    newrow = [CountCounter-1,'=SUM(C2:C'+str(CountCounter+2)+')','=SUM(C2:C'+str(CountCounter+2)+')','=SUM(D2:D'+str(CountCounter+2)+')','=SUM(E2:E'+str(CountCounter+2)+')','=SUM(F2:F'+str(CountCounter+2)+')','=SUM(G2:G'+str(CountCounter+2)+')','=SUM(H2:H'+str(CountCounter+2)+')','=SUM(I2:I'+str(CountCounter+2)+')']
    data = np.vstack([data, newrow])
    CountCounter +=1
    print (data)
    df = pd.DataFrame(data)
    if os.path.exists("C:/Users/msujo/Desktop/Fallout Data/ %s.xlsx" %(datetime.today().strftime('%Y-%m-%d'))) == False:
        df.to_excel("C:/Users/msujo/Desktop/Fallout Data/ %s.xlsx" %(datetime.today().strftime('%Y-%m-%d')),sheet_name="sheet",index = None, header = ['Attempt','Time','Brown','Green', 'Blue','Yellow','Forrest','Red','Total Spawns'])
    else:
        ExcelFile = "C:/Users/msujo/Desktop/Fallout Data/ %s.xlsx" %(datetime.today().strftime('%Y-%m-%d'))
        WorkBook = xl.load_workbook(ExcelFile)
        res = len(WorkBook.sheetnames)                                                           
        count=res+1
        count2 = count 
        with pd.ExcelWriter("C:/Users/msujo/Desktop/Fallout Data/ %s.xlsx" %(datetime.today().strftime('%Y-%m-%d')), engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name="sheet"+str(count),index = None, header = ['Attempt','Time','Brown','Green', 'Blue','Yellow','Forrest','Red','Spawns'])
            #df.to_excel("C:/Users/msujo/Desktop/Fallout Data/ %s.xlsx" %(datetime.today().strftime('%Y-%m-%d')),sheet_name="sheet"+str(count),index = None, header = ['Attempt','Time','Green', 'Blue','Yellow','Forrest','RED'])
    """ExcelFile = "C:/Users/msujo/Desktop/Fallout Data/ %s.xlsx" %(datetime.today().strftime('%Y-%m-%d'))
    WorkBook = xl.load_workbook(ExcelFile)
    sheet = WorkBook.sheetnames
    ws = WorkBook[sheet[count2]]
    p = CountCounter + 3
    cell=ws.cell(row=p, column=2)
    cell = 'FORMULA'
    print (cell) 
    cell = 'FORMULA'"""

def main():
    print("This program is designed to aid in acquiring data on various RNG drops in \nthe video game Fallout 76 ")
    print("")
    print ("Time in seconds since epoch:", get_seconds())
    print("Current Local time", datetime.today().strftime('"%Y-%m-%d %H:%M:%S"'))
    print("")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("INPUT:")
    print("Press 's' to begin a new run.\nPress 'c' every time you leave a world and want to log an attempt. \nPress '2' to end a run.")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("If you successfully find an item in world, use the following inputs to log them.\nMultiple keypresses will log as multiple items.")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("INPUT FOR ITEMS: \nBefore pressing 'c' to log event: \nPress 'g' for Green Find \nPress 'b' for Blue Find \nPress 'y' for Yellow Find \nPress 'f' for Forrest Find \nPress 'r' for Red Find\n\n")

    listener = keyboard.Listener(on_press=on_press)
    listener.start()


                 
    
main()
